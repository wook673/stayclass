# -*- coding: utf-8 -*-
"""정수기 추천 가이드북.xlsx → ments.json (모델코드별 영업 추천멘트) 생성"""
import json
import re
import sys
from pathlib import Path

import openpyxl

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

HERE = Path(__file__).parent
XLSX = Path(r"C:\Users\User\Downloads\정수기 추천 가이드북.xlsx")

# 시트별 (모델/상품명 열, 멘트 열) — 1-based column index
SHEETS = {
    "코웨이":     ("E", "G"),
    "SK매직":     ("C", "E"),
    "쿠쿠":       ("C", "D"),
    "LG정수기":   ("C", "D"),
    "청호나이스": ("C", "E"),
    "현대큐밍":   ("C", "E"),
}

# 모델코드: 영문으로 시작, 숫자를 반드시 포함하는 영숫자/하이픈/언더스코어 조합 (P-2200N 같은 1자 접두 포함)
MODEL_RE = re.compile(r"[A-Z][A-Z0-9_-]*\d[A-Z0-9_-]*")
# 변형 표기: WD722R(K/H/E) → WD722RK, WD722RH, WD722RE (괄호 안 토큰은 1~3자 영숫자)
VARIANT_RE = re.compile(r"([A-Z][A-Z0-9_-]*\d[A-Z0-9_-]*)\(([A-Z0-9]{1,3}(?:/[A-Z0-9]{1,3})*)\)")


def normalize(code: str) -> str:
    """대문자화 + 영숫자만 남기기"""
    return re.sub(r"[^A-Z0-9]", "", code.upper())


def extract_models(text: str) -> list:
    """상품명/모델명 텍스트에서 정규화된 모델코드 목록 추출"""
    text = text.upper()
    codes = []

    def add(c):
        n = normalize(c)
        if n and n not in codes:
            codes.append(n)

    # 1) 변형 표기 전개 (기본 코드도 함께 포함)
    def expand(m):
        base = m.group(1)
        add(base)
        for suffix in m.group(2).split("/"):
            add(base + suffix)
        return " "  # 소비한 부분은 제거

    rest = VARIANT_RE.sub(expand, text)
    # 2) 일반 모델코드
    for m in MODEL_RE.findall(rest):
        add(m)
    return codes


def clean_ment(text: str) -> str:
    """줄바꿈 정리: 연속 개행 → 하나, 각 줄/전체 앞뒤 공백 제거"""
    lines = [ln.strip() for ln in text.splitlines()]
    out, prev_blank = [], False
    for ln in lines:
        if ln:
            out.append(ln)
            prev_blank = False
        elif not prev_blank and out:
            prev_blank = True  # 연속 빈 줄 무시 (빈 줄 자체도 출력에선 제거)
    return "\n".join(out).strip()


def merged_value(ws, cell):
    """셀 값 조회 — 병합 범위 안이면 병합 범위 좌상단 값 사용"""
    if cell.value is not None:
        return cell.value
    for rng in ws.merged_cells.ranges:
        if (rng.min_row <= cell.row <= rng.max_row
                and rng.min_col <= cell.column <= rng.max_col):
            return ws.cell(rng.min_row, rng.min_col).value
    return None


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    entries = []
    stats = {}
    for sheet, (name_col, ment_col) in SHEETS.items():
        if sheet not in wb.sheetnames:
            print(f"경고: 시트 없음 - {sheet}")
            continue
        ws = wb[sheet]
        ni = openpyxl.utils.column_index_from_string(name_col) - 1
        mi = openpyxl.utils.column_index_from_string(ment_col) - 1
        count = 0
        for row in ws.iter_rows():
            name_val = row[ni].value if len(row) > ni else None
            if not name_val:
                continue
            ment_val = merged_value(ws, row[mi]) if len(row) > mi else None
            if not ment_val:
                continue
            models = extract_models(str(name_val))
            if not models:  # 헤더 행 등 모델코드 없는 행은 건너뜀
                continue
            ment = clean_ment(str(ment_val))
            if not ment:
                continue
            entries.append({"models": models, "brand": sheet, "ment": ment})
            count += 1
        stats[sheet] = count

    out = HERE / "ments.json"
    out.write_text(json.dumps(entries, ensure_ascii=False, indent=1), encoding="utf-8")
    total_models = sum(len(e["models"]) for e in entries)
    print(f"OK {out} — 항목 {len(entries)}개, 모델코드 {total_models}개")
    for k, v in stats.items():
        print(f"  {k}: {v}개")


if __name__ == "__main__":
    main()
